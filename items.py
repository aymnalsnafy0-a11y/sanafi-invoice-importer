"""تحميل قائمة الأصناف المرجعية (مُصدَّرة من AccSystem) ومطابقة الأصناف المستخرجة."""

import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz, process

import config
from line_item import ExtractedLine


@dataclass
class ReferenceItem:
    code: str
    name: str
    barcode: str = ""
    default_unit: str = ""
    internal_id: str = ""  # CLS_ID الداخلي بقاعدة البيانات (مطلوب لصيغة الملف القياسي)
    unit_id: str = ""  # UN_ID الداخلي (مطلوب لصيغة الملف القياسي)


def load_item_reference(path: str | Path) -> list[ReferenceItem]:
    """
    يحمّل قائمة الأصناف من ملف مُصدَّر من AccSystem. يدعم صيغتين:
    - اكسل (.xlsx): صف أول كعناوين أعمدة "رقم الصنف" / "اسم الصنف" / "الباركود".
    - ملف قياسي (.AmnC أو .xml): تصدير DataSet داخلي، يحتوي عناصر <CLASSES>
      لكل صنف (CLS_NO=رقم الصنف، CLS_ARNAME=الاسم العربي، GTIN=الباركود).
    """
    path = Path(path)
    if not path.exists():
        return []

    suffix = path.suffix.lower()
    if suffix in (".amnc", ".xml"):
        return _load_from_amnc_xml(path)
    return _load_from_excel(path)


def _load_from_amnc_xml(path: Path) -> list[ReferenceItem]:
    tree = ET.parse(path)
    root = tree.getroot()

    items = []
    for cls_elem in root.findall("CLASSES"):
        code_el = cls_elem.find("CLS_NO")
        name_el = cls_elem.find("CLS_ARNAME")
        enname_el = cls_elem.find("CLS_ENNAME")
        gtin_el = cls_elem.find("GTIN")
        cls_id_el = cls_elem.find("CLS_ID")

        code = code_el.text.strip() if code_el is not None and code_el.text else ""
        name = name_el.text.strip() if name_el is not None and name_el.text else ""
        if not name and enname_el is not None and enname_el.text:
            name = enname_el.text.strip()
        if not code or not name:
            continue

        # رقم الصنف الداخلي بقاعدة البيانات (CLS_ID) - موجود فعلاً بملف
        # "قياسي" المُصدَّر من شاشة الأصناف (تأكدنا من ملف حقيقي 27,395 صنف)،
        # بعكس ما كان مفترضاً سابقاً. بدونه، أي تصدير "ملف قياسي" لاحق من
        # هذا المرجع يفشل بصمت (AccSystem يتجاهل الصف عند CLS_ID=0). لا يوجد
        # حقل مكافئ لـUN_ID (رقم الوحدة الداخلي) بهذا الملف إطلاقاً - يبقى
        # فارغاً هنا؛ التحديث المباشر من قاعدة البيانات هو المصدر الوحيد اللي
        # يوفّره حالياً (راجع db_items.py).
        internal_id = cls_id_el.text.strip() if cls_id_el is not None and cls_id_el.text else ""

        # الصنف الواحد بملف "قياسي" المُصدَّر من شاشة الأصناف ممكن له عدة
        # باركودات - وحدة لكل مستوى تعبئة (حبة/كرتون/بالة...)، بحقول
        # CLS_UN_1_PCODE إلى CLS_UN_6_PCODE (تأكدنا من ملف حقيقي 27,395 صنف:
        # وصلت لـ6 وحدات فعلياً). GTIN فاضي 100% دائماً - نفس نمط اكتشفناه
        # سابقاً بقاعدة البيانات المباشرة. نضيف صف مرجعي منفصل لكل باركود
        # حقيقي موجود، عشان المطابقة تلقاه بغض النظر عن أي وحدة طُبعت
        # بالفاتورة - نفس منطق db_items.py (JOIN مع CLS_UNIT_BARCODE).
        unit_barcodes = []
        for i in range(1, 7):
            pcode_el = cls_elem.find(f"CLS_UN_{i}_PCODE")
            barcode = pcode_el.text.strip() if pcode_el is not None and pcode_el.text else ""
            if not barcode:
                continue
            unit_name_el = cls_elem.find(f"CLS_UN_{i}")
            unit_name = unit_name_el.text.strip() if unit_name_el is not None and unit_name_el.text else ""
            unit_barcodes.append((barcode, unit_name))

        if not unit_barcodes:
            fallback_barcode = gtin_el.text.strip() if gtin_el is not None and gtin_el.text else ""
            unit1_el = cls_elem.find("CLS_UN_1")
            default_unit = unit1_el.text.strip() if unit1_el is not None and unit1_el.text else ""
            items.append(
                ReferenceItem(
                    code=code, name=name, barcode=fallback_barcode, default_unit=default_unit,
                    internal_id=internal_id,
                )
            )
            continue

        for barcode, unit_name in unit_barcodes:
            items.append(
                ReferenceItem(code=code, name=name, barcode=barcode, default_unit=unit_name, internal_id=internal_id)
            )
    return items


def _load_from_excel(path: Path) -> list[ReferenceItem]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ValueError(f"ملف قائمة الأصناف فاضي بدون صف عناوين: {path}")
    col_index = {}
    for idx, header in enumerate(header_row):
        if not header:
            continue
        header = str(header).strip()
        if "رقم" in header and "صنف" in header:
            col_index["code"] = idx
        elif "اسم" in header and "صنف" in header:
            col_index["name"] = idx
        elif "باركود" in header:
            col_index["barcode"] = idx

    if "code" not in col_index or "name" not in col_index:
        raise ValueError(
            "لم يتم العثور على عمودي 'رقم الصنف' و 'اسم الصنف' في ملف قائمة الأصناف. "
            "تأكد من تصدير الملف من AccSystem بالأعمدة الصحيحة."
        )

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[col_index["code"]]
        name = row[col_index["name"]]
        if code is None or name is None:
            continue
        barcode = row[col_index["barcode"]] if "barcode" in col_index else ""
        items.append(
            ReferenceItem(code=str(code).strip(), name=str(name).strip(), barcode=str(barcode or "").strip())
        )
    return items


def _apply_match(line: ExtractedLine, matched: ReferenceItem, score: float) -> None:
    line.match_score = score
    line.needs_review = False
    line.matched_item_code = matched.code
    line.matched_item_name = matched.name
    line.matched_internal_id = matched.internal_id
    line.matched_unit_id = matched.unit_id
    if matched.barcode and not line.barcode:
        line.barcode = matched.barcode
    if matched.default_unit and not line.unit:
        line.unit = matched.default_unit


def match_line_items(lines: list[ExtractedLine], reference: list[ReferenceItem]) -> None:
    """يطابق كل سطر مستخرج مع أقرب صنف في القائمة المرجعية (تعديل في المكان)."""
    if not reference:
        for line in lines:
            line.needs_review = True
        return

    # مطابقة بالباركود أولاً (دقيقة 100% إن وُجدت) قبل التخمين بالاسم -
    # الباركود مُعرِّف فريد بعكس الاسم الذي قد يختلف صياغة عن قاعدتكم
    barcode_lookup = {item.barcode: item for item in reference if item.barcode}

    choices = {i: item.name for i, item in enumerate(reference)}

    for line in lines:
        if line.barcode and line.barcode in barcode_lookup:
            _apply_match(line, barcode_lookup[line.barcode], 100.0)
            continue

        if not line.description:
            continue
        # ملاحظة: WRatio كان يعطي درجات عالية غير موثوقة لمجرد أن الوصف
        # القصير يظهر كجزء من اسم صنف أطول (مثال حقيقي واجهناه: "ثلج" طابق
        # "كيس ثلج النقي 4 كيلو" بدرجة 90 رغم إنه احتمال غلط) - token_sort_ratio
        # لا يكافئ الاحتواء الجزئي بنفس الطريقة، ويبقى يتعامل صح مع اختلاف
        # ترتيب الكلمات (شائع بفواتير الموردين)
        best = process.extractOne(
            line.description, choices, scorer=fuzz.token_sort_ratio
        )
        if best is None:
            continue
        _, score, idx = best
        matched = reference[idx]
        line.match_score = score
        line.needs_review = score < config.AUTO_MATCH_THRESHOLD

        if not line.needs_review:
            # مطابقة واثقة فقط تُستخدم لتعبئة رقم/اسم الصنف - غير ذلك تبقى
            # القراءة الأصلية من الفاتورة كما هي بدل استبدالها باسم/رقم صنف
            # غير مؤكد (قد يكون من فاتورة مورد مختلف تماماً عن أصنافنا)
            _apply_match(line, matched, score)
