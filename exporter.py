"""تصدير بنود الفاتورة المُراجَعة إلى ملف اكسل بصيغة قابلة للاستيراد في AccSystem."""

from pathlib import Path

import openpyxl

import config
from line_item import ExtractedLine

# قواعد مرتبة من الأكثر تحديداً إلى الأعم - يجب فحص "سعر الوحدة" قبل "الوحدة"
# وإلا فإن "وحدة" (كسلسلة فرعية) ستطابق عمود "سعر الوحدة" خطأً.
_COLUMN_RULES = [
    (lambda h: "سعر" in h and "وحدة" in h, lambda l: l.unit_price or ""),
    (lambda h: "رقم" in h and "صنف" in h, lambda l: l.matched_item_code),
    (lambda h: "اسم" in h and "صنف" in h, lambda l: l.matched_item_name or l.description),
    (lambda h: "باركود" in h, lambda l: l.barcode),
    (lambda h: "كمية" in h, lambda l: l.quantity or ""),
    (lambda h: "اجمالي" in h or "إجمالي" in h, lambda l: l.total or ""),
    (lambda h: "وحدة" in h, lambda l: l.unit),
]


def _resolve_field(header: str):
    for predicate, getter in _COLUMN_RULES:
        if predicate(header):
            return getter
    return None


# حماية من "حقن الصيغ" بالاكسل (CSV/Excel Formula Injection): وصف الصنف لصف
# غير مطابَق يجي من قراءة الفاتورة نفسها (مصدر خارجي - مورد قد يكون خبيث)،
# ولو بدأ بأحد هذي الرموز، إكسل يفسّره كصيغة تُنفَّذ تلقائياً عند فتح الملف
# (مثال معروف: =cmd|'/c calc'!A1). نضيف علامة اقتباس بالأول لإجبار إكسل
# يتعامل معه كنص عادي، بدون ما يغيّر القيمة المعروضة فعلياً.
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell_value(value):
    if isinstance(value, str) and value.startswith(_FORMULA_TRIGGER_CHARS):
        return "'" + value
    return value


def export_to_excel(lines: list[ExtractedLine], output_path: str | Path) -> Path:
    """
    يكتب البنود إلى ملف اكسل. إن وُجد قالب حقيقي في config.IMPORT_TEMPLATE_PATH
    (مُصدَّر من AccSystem عبر "إنشاء نموذج اكسل")، يُستخدم صف عناوينه بالضبط.
    وإلا، تُستخدم الأعمدة الافتراضية في config.DEFAULT_OUTPUT_COLUMNS.
    """
    output_path = Path(output_path)
    template_path = Path(config.IMPORT_TEMPLATE_PATH)

    if template_path.exists():
        template_wb = openpyxl.load_workbook(template_path, data_only=True)
        headers = [c.value for c in next(template_wb.active.iter_rows(min_row=1, max_row=1))]
    else:
        headers = config.DEFAULT_OUTPUT_COLUMNS

    getters = [_resolve_field(str(h)) if h else None for h in headers]

    # يُبنى ملف جديد بنفس صف العناوين فقط (بدون أي صفوف بيانات قديمة/تجريبية
    # قد تكون موجودة في ملف القالب المصدَّر من النظام)، ثم تُضاف بنود الفاتورة.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    start_row = 2

    for row_offset, line in enumerate(lines):
        row_num = start_row + row_offset
        for col_num, getter in enumerate(getters, start=1):
            value = getter(line) if getter else ""
            ws.cell(row=row_num, column=col_num, value=_sanitize_cell_value(value))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
