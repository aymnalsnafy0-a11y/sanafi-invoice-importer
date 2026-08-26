"""
Benchmark مستقل (بدون أي ربط بواجهة Tkinter/app.py) لقياس أداء مسار
الاستخراج + المطابقة المحلية الحقيقي على فاتورة حقيقية سبق تأكيدها.

يعيد استخدام دوال الإنتاج الفعلية (items.match_line_items،
matching_engine.suggest_candidates/enhance_one) بنفس التسلسل المستخدم في
app.py::_extract_one_invoice بالضبط - بدون استيراد app.py أو Tkinter إطلاقاً.

الاستخدام:
    python tools/benchmark_invoices.py --folder test_invoices

المدخلات المطلوبة داخل --folder (تُكتشف بالامتداد فقط):
    *.pdf   - الفاتورة الخام (مدخل Vision)
    *.AmnC  - كتالوج الأصناف المرجعي (يُطابَق ضده)
    *.xlsx  - تصدير Excel سابق مؤكَّد لنفس الفاتورة (Ground Truth الأساسي)
    *.Amn   - تصدير .Amn سابق مؤكَّد لنفس الفاتورة (مرجع تقاطع/تحقق)

الوضع الافتراضي: LOCAL ONLY (بدون Semantic AI)، بدون تعديل أي state حقيقي،
بدون أي اتصال Oracle.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
import tempfile
import time
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path

APP_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(APP_ROOT))

import openpyxl
from rapidfuzz import fuzz

_LINE_ALIGN_MIN_SCORE = 45  # عن قصد أقل من عتبة القبول (80%) - هذا فقط لإيجاد
# أي سطر فاتورة يقابل أي سطر Ground Truth، مو حكم ثقة على صحة المطابقة.


# ==================== نماذج البيانات ====================

@dataclass
class GroundTruthLine:
    code: str
    name: str
    barcode: str
    unit: str
    quantity: float | None
    unit_price: float | None
    total: float | None
    source: str  # "excel" | "amn"


@dataclass
class DiscoveredFiles:
    pdf_path: Path
    amnc_path: Path
    xlsx_path: Path
    amn_path: Path
    skipped_pdfs: list[Path]


@dataclass
class PerLineRecord:
    row_type: str  # "paired" | "missing_ground_truth" | "extra_extracted"
    align_method: str | None = None  # "barcode_exact" | "fuzzy_description"
    align_score: float | None = None
    extracted_index: int | None = None
    truth_index: int | None = None

    extracted_description: str = ""
    extracted_barcode_original: str = ""
    extracted_barcode_final: str = ""
    extracted_unit: str = ""
    extracted_qty: float | None = None
    extracted_price: float | None = None
    extracted_total: float | None = None

    matched_item_code: str = ""
    matched_item_name: str = ""
    needs_review: bool | None = None
    match_score: float | None = None
    match_reason: str = ""
    was_barcode_shortcut: bool = False

    top1_local_code: str | None = None
    top1_local_name: str | None = None
    top1_local_confidence: float | None = None
    top1_local_reason: str | None = None
    top1_matches_expected: bool | None = None

    expected_code: str = ""
    expected_name: str = ""
    expected_barcode: str = ""
    expected_qty: float | None = None
    expected_price: float | None = None
    expected_total: float | None = None

    verdict: str | None = None  # "correct" | "wrong" | "unresolved" | None
    is_false_auto_accept: bool = False
    is_wrong_but_flagged: bool = False
    barcode_match: bool | None = None
    qty_match: bool | None = None
    price_match: bool | None = None


# ==================== أدوات مساعدة عامة ====================

def _fail(msg: str) -> None:
    print(f"[ERROR] {msg}", file=sys.stderr)
    sys.exit(2)


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _floats_close(a: float | None, b: float | None, tol: float = 0.01) -> bool | None:
    if a is None or b is None:
        return None
    return abs(a - b) <= tol


# ==================== 1) اكتشاف الملفات ====================

def _require_exactly_one(matches: list[Path], folder: Path, pattern: str, label: str) -> Path:
    if not matches:
        _fail(f"لا يوجد {pattern} ({label}) داخل {folder}. مطلوب ملف واحد بالضبط.")
    if len(matches) > 1:
        names = ", ".join(m.name for m in matches)
        _fail(f"أكثر من {pattern} داخل {folder} ({names}) - مطلوب ملف واحد بالضبط، بدون تخمين. احذف الزائد.")
    return matches[0]


def discover_input_files(folder: Path, limit: int) -> DiscoveredFiles:
    if limit != 1:
        print(f"[WARN] --limit={limit} مطلوب، لكن وضع تعدد الفواتير غير مُنفَّذ بعد - سيُعالَج فاتورة واحدة بالضبط بغض النظر.")

    if not folder.is_dir():
        _fail(f"المجلد غير موجود: {folder}")

    pdfs = sorted(folder.glob("*.pdf"))
    amncs = sorted(folder.glob("*.AmnC"))
    xlsxs = sorted(folder.glob("*.xlsx"))
    amns = sorted(folder.glob("*.Amn"))

    if not pdfs:
        _fail(f"لا يوجد *.pdf داخل {folder}. ضع فيه الفاتورة الخام.")
    skipped = pdfs[1:]
    if skipped:
        print(f"[WARN] وُجد {len(pdfs)} ملف PDF، سيُستخدم {pdfs[0].name}؛ تم تجاوز: {', '.join(p.name for p in skipped)}")

    amnc = _require_exactly_one(amncs, folder, "*.AmnC", "كتالوج الأصناف المرجعي")
    xlsx = _require_exactly_one(xlsxs, folder, "*.xlsx", "تصدير Excel مؤكَّد")
    amn = _require_exactly_one(amns, folder, "*.Amn", "تصدير .Amn مؤكَّد")

    # حماية دفاعية - أبداً ما نستبدل صامتين بملفات المشروع الجذرية (fixture/template)
    forbidden_pairs = [(amnc, APP_ROOT / "item_master.AmnC"), (xlsx, APP_ROOT / "aman_import_template.xlsx")]
    for p, forbidden in forbidden_pairs:
        if p.resolve() == forbidden.resolve():
            _fail(f"{p} هو نفسه ملف المشروع الجذري ({forbidden.name}) - رفض استخدامه كـGround Truth. ضع تصدير حقيقي خاص بهذه الفاتورة داخل {folder}.")

    return DiscoveredFiles(pdfs[0], amnc, xlsx, amn, skipped)


# ==================== 2) قراءة Ground Truth ====================

def _classify_header(h: str) -> str | None:
    # مطابِق (اتجاه قراءة) لقواعد exporter.py::_COLUMN_RULES (اتجاه كتابة) -
    # يجب إبقاؤه متزامناً معها يدوياً لو تغيّرت هناك. الترتيب مهم: "سعر"+"وحدة"
    # يُفحص قبل "وحدة" المجردة لأن "سعر الوحدة" تحتوي "وحدة" كـsubstring.
    if "سعر" in h and "وحدة" in h:
        return "unit_price"
    if "رقم" in h and "صنف" in h:
        return "code"
    if "اسم" in h and "صنف" in h:
        return "name"
    if "باركود" in h:
        return "barcode"
    if "كمية" in h:
        return "quantity"
    if "اجمالي" in h or "إجمالي" in h:
        return "total"
    if "وحدة" in h:
        return "unit"
    return None


def _strip_formula_guard(v):
    # exporter.py::_sanitize_cell_value يضيف "'" بادئة لمنع حقن الصيغ
    if isinstance(v, str) and v.startswith("'"):
        return v[1:]
    return v


def load_ground_truth_excel(path: Path) -> list[GroundTruthLine]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        _fail(f"{path} بلا صف عناوين.")

    col_index: dict[str, int] = {}
    for idx, header in enumerate(header_row):
        if not header:
            continue
        field_name = _classify_header(str(header).strip())
        if field_name and field_name not in col_index:
            col_index[field_name] = idx
    if "code" not in col_index or "name" not in col_index:
        _fail(f"{path}: تعذّر إيجاد عمودي 'رقم الصنف'/'اسم الصنف'.")

    def cell(row, key):
        return _strip_formula_guard(row[col_index[key]]) if key in col_index else None

    out: list[GroundTruthLine] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(cell(row, "code") or "").strip()
        name = str(cell(row, "name") or "").strip()
        if not code and not name:
            continue  # صف فاضي بنهاية الملف
        out.append(GroundTruthLine(
            code=code, name=name,
            barcode=str(cell(row, "barcode") or "").strip(),
            unit=str(cell(row, "unit") or "").strip(),
            quantity=_to_float(cell(row, "quantity")),
            unit_price=_to_float(cell(row, "unit_price")),
            total=_to_float(cell(row, "total")),
            source="excel",
        ))
    return out


def load_ground_truth_amn(path: Path) -> list[GroundTruthLine]:
    root = ET.parse(path).getroot()
    out: list[GroundTruthLine] = []
    for det in root.findall("DOC_DET_TABLE"):
        def text(tag):
            el = det.find(tag)
            return el.text.strip() if el is not None and el.text else ""

        name = text("CLS_ARNAME") or text("CLS_ENNAME")
        out.append(GroundTruthLine(
            code=text("CLS_NO"), name=name,
            barcode=text("DOC_D_CLS_BARCODE"),
            unit=text("DOC_D_UN_NAME"),
            quantity=_to_float(text("DOC_D_QLT") or None),
            unit_price=_to_float(text("DOC_D_PRICE") or None),
            total=_to_float(text("DOC_D_FORIGNVALUE") or None),
            source="amn",
        ))
    return out


def cross_check_ground_truth(excel: list[GroundTruthLine], amn: list[GroundTruthLine]) -> dict:
    discrepancies = []
    if len(excel) != len(amn):
        discrepancies.append({"type": "line_count_mismatch", "excel_count": len(excel), "amn_count": len(amn)})

    for i in range(min(len(excel), len(amn))):
        e, a = excel[i], amn[i]
        for field_name, tol in (("code", None), ("barcode", None), ("quantity", 0.01), ("unit_price", 0.01)):
            ev, av = getattr(e, field_name), getattr(a, field_name)
            if tol is not None and ev is not None and av is not None:
                mismatched = abs(ev - av) > tol
            else:
                mismatched = ev != av
            if mismatched:
                discrepancies.append({"type": "field_mismatch", "index": i, "field": field_name, "excel_value": ev, "amn_value": av})

    reliable = len(excel) == len(amn)
    return {
        "discrepancies": discrepancies,
        "reliable": reliable,
        "excel_line_count": len(excel),
        "amn_line_count": len(amn),
    }


# ==================== 3) فحص بنية AmnC (استكشاف فقط - بدون دمج) ====================

def survey_amnc_structure(path: Path) -> dict:
    root = ET.parse(path).getroot()
    classes = root.findall("CLASSES")
    total = len(classes)
    multi_unit = any_barcode = gtin_only = trans_populated = 0

    for cls in classes:
        pcodes = []
        for i in range(1, 7):
            el = cls.find(f"CLS_UN_{i}_PCODE")
            if el is not None and el.text and el.text.strip():
                pcodes.append(i)
        if len(pcodes) > 1:
            multi_unit += 1
        gtin_el = cls.find("GTIN")
        gtin_val = gtin_el.text.strip() if gtin_el is not None and gtin_el.text else ""
        if pcodes:
            any_barcode += 1
        elif gtin_val:
            gtin_only += 1
        has_trans = False
        for i in range(2, 7):
            el = cls.find(f"CLS_UN_{i}_TRANS")
            if el is not None and el.text and el.text.strip() not in ("", "0"):
                has_trans = True
                break
        if has_trans:
            trans_populated += 1

    ub_rows = root.findall("CLS_UNIT_BARCODE")
    class_ids = set()
    for cls in classes:
        t = cls.find("CLS_ID")
        if t is not None and t.text:
            class_ids.add(t.text.strip())

    sample_n = min(50, len(ub_rows))
    sample_hits = 0
    for row in ub_rows[:sample_n]:
        t = row.find("CLS_ID")
        if t is not None and t.text and t.text.strip() in class_ids:
            sample_hits += 1

    def pct(n):
        return round(100 * n / total, 1) if total else None

    return {
        "total_classes": total,
        "multi_unit_barcode_items": {"count": multi_unit, "pct": pct(multi_unit)},
        "items_with_any_unit_barcode": {"count": any_barcode, "pct": pct(any_barcode)},
        "items_relying_on_gtin_only": {"count": gtin_only, "pct": pct(gtin_only)},
        "items_with_populated_trans_factor": {"count": trans_populated, "pct": pct(trans_populated)},
        "cls_unit_barcode_table": {
            "row_count": len(ub_rows),
            "spot_check_sample_size": sample_n,
            "spot_check_cls_id_join_matches": sample_hits,
            "spot_check_join_match_rate": round(sample_hits / sample_n, 3) if sample_n else None,
        },
        "note": (
            "فحص للقراءة فقط - غير مربوط بـmatching_engine.py أو items.py إطلاقاً. "
            "حقول قد تفيد لاحقاً (المعنى مُستنتَج من اسم الحقل فقط، غير مؤكَّد مقابل "
            "توثيق AccSystem الرسمي): CLS_ORD_PUR, CLS_IS_DEF, CLS_NO_VND, "
            "CLS_UNIT_BARCODE.CLS_BAR_PRICE, CLS_UNIT_BARCODE.UN_ID."
        ),
    }


# ==================== 4) عزل ملفات الحالة (بدون تعديل أي شيء حقيقي) ====================

def isolate_state_files(config, learned_matches, settings_module, semantic_matcher, usage_tracker,
                         seed_learned_matches: bool) -> Path:
    tmp_dir = Path(tempfile.mkdtemp(prefix="benchmark_invoices_state_"))

    real_learned = config.BASE_DIR / "learned_matches.json"
    temp_learned = tmp_dir / "learned_matches.json"
    if seed_learned_matches and real_learned.exists():
        temp_learned.write_text(real_learned.read_text(encoding="utf-8"), encoding="utf-8")
    learned_matches._MATCHES_FILE = temp_learned

    # matching_settings.json - يُنسَخ (قراءة فقط) من الملف الحقيقي لو موجود،
    # عشان العتبات (auto_accept_threshold, min_confidence_gap) تعكس الإعداد
    # الفعلي الحالي - هذا إعداد خوارزمية، مو تاريخ تأكيدات مستخدم
    real_settings = config.BASE_DIR / "matching_settings.json"
    temp_settings = tmp_dir / "matching_settings.json"
    if real_settings.exists():
        temp_settings.write_text(real_settings.read_text(encoding="utf-8"), encoding="utf-8")
    settings_module._SETTINGS_FILE = temp_settings

    semantic_matcher._CACHE_FILE = tmp_dir / "semantic_rerank_cache.json"
    usage_tracker._USAGE_FILE = tmp_dir / "usage_state.json"

    return tmp_dir


# ==================== 5) محاذاة الأسطر ====================

def align_lines(extracted: list, original_barcodes: list[str], ground_truth: list[GroundTruthLine]):
    unpaired_truth_idx = list(range(len(ground_truth)))

    gt_by_barcode: dict[str, list[int]] = {}
    for j in unpaired_truth_idx:
        bc = ground_truth[j].barcode.strip()
        if bc:
            gt_by_barcode.setdefault(bc, []).append(j)

    pairs: list[tuple[int, int, str, float]] = []
    still_unpaired_extracted = []
    for i, _line in enumerate(extracted):
        bc = (original_barcodes[i] or "").strip()
        bucket = gt_by_barcode.get(bc) if bc else None
        if bucket:
            j = bucket.pop(0)
            pairs.append((i, j, "barcode_exact", 100.0))
        else:
            still_unpaired_extracted.append(i)

    paired_truth = {j for _, j, _, _ in pairs}
    still_unpaired_truth = [j for j in unpaired_truth_idx if j not in paired_truth]

    scored = []
    for i in still_unpaired_extracted:
        desc = extracted[i].description
        if not desc:
            continue
        for j in still_unpaired_truth:
            name = ground_truth[j].name
            if not name:
                continue
            score = fuzz.token_sort_ratio(desc, name)
            if score >= _LINE_ALIGN_MIN_SCORE:
                scored.append((score, i, j))
    scored.sort(key=lambda t: t[0], reverse=True)

    used_i, used_j = set(), set()
    for score, i, j in scored:
        if i in used_i or j in used_j:
            continue
        pairs.append((i, j, "fuzzy_description", float(score)))
        used_i.add(i)
        used_j.add(j)

    paired_i = {p[0] for p in pairs}
    paired_j = {p[1] for p in pairs}
    missing_truth = [j for j in range(len(ground_truth)) if j not in paired_j]
    extra_extracted = [i for i in range(len(extracted)) if i not in paired_i]
    return pairs, missing_truth, extra_extracted


# ==================== 6) بناء سجلات الأسطر + الحكم لكل سطر ====================

def build_per_line_records(pairs, missing_truth_idx, extra_extracted_idx, extracted: list,
                            ground_truth: list[GroundTruthLine], original_barcodes: list[str],
                            top1_by_index: dict, barcode_shortcut: list[bool]) -> list[PerLineRecord]:
    records: list[PerLineRecord] = []

    for i, j, method, score in pairs:
        line = extracted[i]
        gt = ground_truth[j]

        code = (line.matched_item_code or "").strip()
        expected_code = gt.code.strip()
        if not code:
            verdict = "unresolved"
        elif code == expected_code:
            verdict = "correct"
        else:
            verdict = "wrong"

        is_false_auto_accept = (line.needs_review is False) and verdict == "wrong"
        is_wrong_but_flagged = verdict == "wrong" and line.needs_review is True

        top1 = top1_by_index.get(i)
        top1_code = top1.item.code if top1 else None

        rec = PerLineRecord(
            row_type="paired", align_method=method, align_score=score,
            extracted_index=i, truth_index=j,
            extracted_description=line.description,
            extracted_barcode_original=(original_barcodes[i] or ""),
            extracted_barcode_final=line.barcode,
            extracted_unit=line.unit,
            extracted_qty=line.quantity,
            extracted_price=line.unit_price,
            extracted_total=line.total,
            matched_item_code=code,
            matched_item_name=line.matched_item_name,
            needs_review=line.needs_review,
            match_score=line.match_score,
            match_reason=line.match_reason,
            was_barcode_shortcut=barcode_shortcut[i],
            top1_local_code=top1_code,
            top1_local_name=(top1.item.name if top1 else None),
            top1_local_confidence=(top1.confidence if top1 else None),
            top1_local_reason=(top1.reason if top1 else None),
            top1_matches_expected=(top1_code == expected_code) if top1 else None,
            expected_code=expected_code, expected_name=gt.name, expected_barcode=gt.barcode,
            expected_qty=gt.quantity, expected_price=gt.unit_price, expected_total=gt.total,
            verdict=verdict, is_false_auto_accept=is_false_auto_accept, is_wrong_but_flagged=is_wrong_but_flagged,
            barcode_match=((original_barcodes[i] or "").strip() == gt.barcode.strip()) if gt.barcode else None,
            qty_match=_floats_close(line.quantity, gt.quantity),
            price_match=_floats_close(line.unit_price, gt.unit_price),
        )
        records.append(rec)

    for j in missing_truth_idx:
        gt = ground_truth[j]
        records.append(PerLineRecord(
            row_type="missing_ground_truth", truth_index=j,
            expected_code=gt.code, expected_name=gt.name, expected_barcode=gt.barcode,
            expected_qty=gt.quantity, expected_price=gt.unit_price, expected_total=gt.total,
        ))

    for i in extra_extracted_idx:
        line = extracted[i]
        records.append(PerLineRecord(
            row_type="extra_extracted", extracted_index=i,
            extracted_description=line.description,
            extracted_barcode_original=(original_barcodes[i] or ""),
            extracted_barcode_final=line.barcode,
            extracted_unit=line.unit, extracted_qty=line.quantity,
            extracted_price=line.unit_price, extracted_total=line.total,
            matched_item_code=line.matched_item_code, matched_item_name=line.matched_item_name,
            needs_review=line.needs_review, match_score=line.match_score, match_reason=line.match_reason,
            was_barcode_shortcut=barcode_shortcut[i],
        ))

    return records


# ==================== 7) المقاييس المجمّعة ====================

def compute_metrics(records: list[PerLineRecord]) -> dict:
    paired = [r for r in records if r.row_type == "paired"]
    n_missing = sum(1 for r in records if r.row_type == "missing_ground_truth")
    n_extra = sum(1 for r in records if r.row_type == "extra_extracted")
    n_expected = len(paired) + n_missing
    n_extracted = len(paired) + n_extra

    correct = [r for r in paired if r.verdict == "correct"]
    wrong = [r for r in paired if r.verdict == "wrong"]
    unresolved = [r for r in paired if r.verdict == "unresolved"]
    false_auto = [r for r in paired if r.is_false_auto_accept]
    wrong_flagged = [r for r in paired if r.is_wrong_but_flagged]
    needs_review = [r for r in paired if r.needs_review]

    bc_expected = [r for r in paired if r.expected_barcode]
    bc_present = [r for r in bc_expected if r.extracted_barcode_original]
    bc_correct = [r for r in bc_present if r.extracted_barcode_original.strip() == r.expected_barcode.strip()]

    enhanced = [r for r in paired if not r.was_barcode_shortcut]
    top1_correct = [r for r in enhanced if r.top1_matches_expected]

    auto_accepted = [r for r in paired if r.needs_review is False]
    auto_accept_correct = [r for r in auto_accepted if r.verdict == "correct"]

    def comparable(attr_expected, attr_actual):
        pairs_ = [(getattr(r, attr_expected), getattr(r, attr_actual)) for r in paired]
        comp = [(e, a) for e, a in pairs_ if e is not None and a is not None]
        matches = sum(1 for e, a in comp if abs(e - a) <= 0.01)
        return len(comp), matches, len(pairs_) - len(comp)

    qty_n, qty_ok, qty_incomparable = comparable("expected_qty", "extracted_qty")
    price_n, price_ok, price_incomparable = comparable("expected_price", "extracted_price")

    return {
        "expected_line_count": n_expected,
        "extracted_line_count": n_extracted,
        "paired_count": len(paired),
        "missing_ground_truth_count": n_missing,
        "extra_extracted_count": n_extra,
        "line_recall": round(len(paired) / n_expected, 4) if n_expected else None,

        "barcode_expected_count": len(bc_expected),
        "barcode_extracted_present_count": len(bc_present),
        "barcode_presence_rate": round(len(bc_present) / len(bc_expected), 4) if bc_expected else None,
        "barcode_accuracy_when_present": round(len(bc_correct) / len(bc_present), 4) if bc_present else None,
        "barcode_accuracy_over_expected": round(len(bc_correct) / len(bc_expected), 4) if bc_expected else None,

        "correct_count": len(correct), "wrong_count": len(wrong), "unresolved_count": len(unresolved),
        "item_code_accuracy_over_paired": round(len(correct) / len(paired), 4) if paired else None,
        "item_code_accuracy_over_expected": round(len(correct) / n_expected, 4) if n_expected else None,

        "top1_local_match_denominator": len(enhanced),
        "top1_local_match_accuracy": round(len(top1_correct) / len(enhanced), 4) if enhanced else None,

        "auto_accepted_count": len(auto_accepted),
        "auto_accept_accuracy": round(len(auto_accept_correct) / len(auto_accepted), 4) if auto_accepted else None,

        "false_auto_accept_count": len(false_auto),
        "wrong_but_flagged_count": len(wrong_flagged),
        "needs_review_count": len(needs_review),

        "qty_comparable_count": qty_n, "qty_correct_count": qty_ok, "qty_incomparable_count": qty_incomparable,
        "qty_accuracy": round(qty_ok / qty_n, 4) if qty_n else None,
        "price_comparable_count": price_n, "price_correct_count": price_ok, "price_incomparable_count": price_incomparable,
        "price_accuracy": round(price_ok / price_n, 4) if price_n else None,
    }


# ==================== 8) كاش Vision (--skip-vision) ====================

def vision_cache_path(folder: Path, pdf_stem: str) -> Path:
    return folder / ".benchmark_cache" / f"{pdf_stem}.vision_raw.json"


def save_vision_cache(path: Path, page_results: list) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = [
        {
            "items": [asdict(line) for line in pr.items],
            "subtotal_before_tax": pr.subtotal_before_tax,
            "tax_amount": pr.tax_amount,
            "grand_total_with_tax": pr.grand_total_with_tax,
            "supplier_name": pr.supplier_name,
        }
        for pr in page_results
    ]
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_vision_cache(path: Path, VisionPageResult, ExtractedLine):
    if not path.exists():
        _fail(f"--skip-vision مطلوب لكن لا يوجد كاش عند {path}. شغّل مرة واحدة بدون --skip-vision أولاً - لا يوجد رجوع صامت لنداء Vision حقيقي.")
    raw = json.loads(path.read_text(encoding="utf-8"))
    return [
        VisionPageResult(
            items=[ExtractedLine(**item) for item in pr["items"]],
            subtotal_before_tax=pr.get("subtotal_before_tax"),
            tax_amount=pr.get("tax_amount"),
            grand_total_with_tax=pr.get("grand_total_with_tax"),
            supplier_name=pr.get("supplier_name"),
        )
        for pr in raw
    ]


# ==================== 9) الإخراج ====================

def write_json_report(path: Path, args, started, files: DiscoveredFiles, timing: dict,
                       api_cost_this_run: float, gt_cross_check: dict, amnc_survey: dict,
                       records: list[PerLineRecord], metrics: dict, tmp_state_dir: Path) -> None:
    payload = {
        "run_metadata": {
            "timestamp_utc": started.isoformat(),
            "flags": {
                "folder": str(args.folder), "limit": args.limit,
                "with_semantic": args.with_semantic, "skip_vision": args.skip_vision,
                "seed_learned_matches": args.seed_learned_matches,
            },
            "discovered_files": {
                "pdf": str(files.pdf_path), "amnc": str(files.amnc_path),
                "xlsx": str(files.xlsx_path), "amn": str(files.amn_path),
                "skipped_pdfs": [str(p) for p in files.skipped_pdfs],
            },
            "isolated_state_dir": str(tmp_state_dir),
            "api_cost_this_run_usd": api_cost_this_run,
        },
        "timing_seconds": timing,
        "ground_truth_cross_check": gt_cross_check,
        "amnc_structure_survey": amnc_survey,
        "lines": [asdict(r) for r in records],
        "metrics": metrics,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _suspicious_reasons(r: PerLineRecord) -> list[str]:
    """أسباب صريحة لماذا سطر يستاهل مراجعة بشرية/مراجع مستقل (Gemini) -
    مبنية على حقول PerLineRecord الموجودة فعلاً، بدون أي منطق مطابقة جديد."""
    reasons = []
    if r.row_type == "missing_ground_truth":
        reasons.append("لم يُستخرَج إطلاقاً (موجود بـGround Truth، غائب عن نتيجة الاستخراج)")
        return reasons
    if r.row_type == "extra_extracted":
        reasons.append("استُخرِج سطر إضافي لا يقابله شيء بـGround Truth (احتمال تكرار/اختراع)")
        return reasons
    if r.is_false_auto_accept:
        reasons.append("⚠⚠ قبول تلقائي خاطئ - أخطر حالة ممكنة (needs_review=False لكن الكود غلط)")
    if r.verdict == "wrong" and not r.is_false_auto_accept:
        reasons.append("مطابقة خاطئة (لحسن الحظ أُحيلت لمراجعة بشرية)")
    if r.verdict == "unresolved":
        reasons.append("لم يُحسَم - احتاج مراجعة بشرية (matched_item_code فاضي)")
    if r.align_method == "fuzzy_description" and (r.align_score or 0) < 70:
        reasons.append(f"محاذاة الاستخراج بـGround Truth ضعيفة نسبياً (تشابه اسم {r.align_score:.0f}% فقط) - قد تكون المقارنة نفسها غير دقيقة")
    if r.barcode_match is False:
        reasons.append("الباركود المستخرَج لا يطابق باركود Ground Truth")
    if r.qty_match is False:
        reasons.append("الكمية المستخرَجة لا تطابق Ground Truth")
    if r.price_match is False:
        reasons.append("سعر الوحدة المستخرَج لا يطابق Ground Truth")
    if r.top1_local_reason and "⚠" in r.top1_local_reason:
        reasons.append("المرشّح المحلي الأفضل عنده تعارض بنيوي محتمل (حجم/عدد/وحدة) بحسب سبب المحرك المحلي")
    return reasons


def build_review_package(records: list[PerLineRecord], metrics: dict, amnc_survey: dict,
                          gt_cross_check: dict, args, files: DiscoveredFiles) -> dict:
    """حزمة مراجعة مختصرة لمراجع مستقل (مثلاً Gemini) - نفس بيانات
    benchmark_results.json بس مرشَّحة لسطور "تستاهل انتباه" بس + المقاييس
    المجمّعة، بدون أي أسرار/بيانات اعتماد. لا تعديل مطابقة هنا - عرض فقط."""
    flagged = []
    for r in records:
        reasons = _suspicious_reasons(r)
        if reasons:
            entry = asdict(r)
            entry["suspicious_reasons"] = reasons
            flagged.append(entry)

    return {
        "source_invoice": files.pdf_path.name,
        "with_semantic": args.with_semantic,
        "metrics": metrics,
        "amnc_structure_survey": amnc_survey,
        "ground_truth_cross_check": gt_cross_check,
        "flagged_line_count": len(flagged),
        "total_line_count": len(records),
        "flagged_lines": flagged,
    }


def write_review_package(path: Path, records: list[PerLineRecord], metrics: dict, amnc_survey: dict,
                          gt_cross_check: dict, args, files: DiscoveredFiles) -> None:
    package = build_review_package(records, metrics, amnc_survey, gt_cross_check, args, files)
    path.write_text(json.dumps(package, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv_summary(path: Path, records: list[PerLineRecord]) -> None:
    if not records:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(asdict(records[0]).keys())
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in records:
            writer.writerow(asdict(r))


def print_console_summary(metrics: dict, amnc_survey: dict, gt_cross_check: dict,
                           api_cost_this_run: float, timing: dict) -> None:
    print("\n" + "=" * 60)
    print("!!! FALSE AUTO ACCEPTS:", metrics["false_auto_accept_count"], "!!!")
    print(f"Ground truth reliable (Excel/.Amn agree): {gt_cross_check['reliable']}")
    print("=" * 60)
    print(f"expected_line_count={metrics['expected_line_count']}  extracted_line_count={metrics['extracted_line_count']}")
    print(f"line_recall={metrics['line_recall']}")
    print(f"correct={metrics['correct_count']}  wrong={metrics['wrong_count']}  unresolved={metrics['unresolved_count']}")
    print(f"wrong_but_flagged={metrics['wrong_but_flagged_count']}  needs_review_count={metrics['needs_review_count']}")
    print(f"top1_local_match_accuracy={metrics['top1_local_match_accuracy']}  auto_accept_accuracy={metrics['auto_accept_accuracy']}")
    print(f"barcode_presence_rate={metrics['barcode_presence_rate']}  barcode_accuracy_when_present={metrics['barcode_accuracy_when_present']}")
    print(f"qty_accuracy={metrics['qty_accuracy']}  price_accuracy={metrics['price_accuracy']}")
    print(f"\napi_cost_this_run_usd={api_cost_this_run}")
    print(f"timing={json.dumps(timing, ensure_ascii=False)}")
    print(f"\namnc total_classes={amnc_survey['total_classes']}  multi_unit={amnc_survey['multi_unit_barcode_items']}")
    print("=" * 60)


# ==================== CLI ====================

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Benchmark مستقل (LOCAL ONLY افتراضياً) لمسار الاستخراج+المطابقة الحقيقي.")
    p.add_argument("--folder", type=Path, default=Path("test_invoices"))
    p.add_argument("--limit", type=int, default=1, help="أقصى عدد فواتير. فقط 1 مُنفَّذة/مُختبَرة اليوم.")
    p.add_argument("--with-semantic", action="store_true", default=False,
                    help="تكلفة إضافية حقيقية: يفعّل نداءات Semantic AI شبكية حقيقية (allow_semantic=True). معطّل افتراضياً - لا تستخدمه بالتشغيل الأول.")
    p.add_argument("--skip-vision", action="store_true", default=False,
                    help="استخدام كاش Vision محفوظ مسبقاً بدل نداء API حقيقي. يفشل بوضوح لو ما فيه كاش.")
    p.add_argument("--seed-learned-matches", action="store_true", default=False,
                    help="ينسخ (قراءة فقط) learned_matches.json الحقيقي للنسخة المعزولة. معطّل افتراضياً (بداية باردة).")
    return p


def main(argv=None):
    args = build_arg_parser().parse_args(argv)
    started = datetime.now(timezone.utc)
    timing: dict = {}
    folder = args.folder.resolve()

    files = discover_input_files(folder, args.limit)

    import config
    import learned_matches
    import settings as settings_module
    import semantic_matcher
    import usage_tracker
    import matching_engine
    from items import load_item_reference, match_line_items
    from line_item import ExtractedLine
    from pdf_utils import load_pages_as_images
    from vision_extract import extract_line_items_vision, VisionPageResult

    real_ai_enabled = usage_tracker.is_ai_enabled()
    if not real_ai_enabled:
        print("[WARN] الذكاء الاصطناعي معطّل بالواجهة الحقيقية (ai_enabled=false) - هذا التشغيل سيستدعي API الحقيقي رغم ذلك لأنك استدعيته صراحة عبر CLI.")

    tmp_state_dir = isolate_state_files(config, learned_matches, settings_module, semantic_matcher, usage_tracker,
                                         seed_learned_matches=args.seed_learned_matches)

    reference = load_item_reference(files.amnc_path)
    if not reference:
        _fail(f"{files.amnc_path} حمّل صفر أصناف مرجعية - تحقق من الملف.")
    ref_index = matching_engine.build_reference_attrs_index(reference)

    amnc_survey = survey_amnc_structure(files.amnc_path)

    gt_excel = load_ground_truth_excel(files.xlsx_path)
    gt_amn = load_ground_truth_amn(files.amn_path)
    gt_cross_check = cross_check_ground_truth(gt_excel, gt_amn)
    ground_truth = gt_excel
    if not gt_cross_check["reliable"]:
        print(f"[WARN] Ground Truth غير موثوق: Excel فيه {len(gt_excel)} سطر، .Amn فيه {len(gt_amn)} - سيُقيَّم مقابل Excel رغم ذلك، راجع تقرير JSON.")

    cache_path = vision_cache_path(folder, files.pdf_path.stem)

    t0 = time.perf_counter()
    pages = load_pages_as_images(files.pdf_path)
    timing["pdf_page_load"] = time.perf_counter() - t0

    if args.skip_vision:
        page_results = load_vision_cache(cache_path, VisionPageResult, ExtractedLine)
        timing["vision_total"] = None
    else:
        t0 = time.perf_counter()
        page_results = [extract_line_items_vision(img) for img in pages]
        timing["vision_total"] = time.perf_counter() - t0
        save_vision_cache(cache_path, page_results)

    vision_lines: list = []
    subtotal_before_tax = tax_amount = grand_total_with_tax = supplier_name = None
    for pr in page_results:
        vision_lines.extend(pr.items)
        if subtotal_before_tax is None:
            subtotal_before_tax = pr.subtotal_before_tax
        if tax_amount is None:
            tax_amount = pr.tax_amount
        if grand_total_with_tax is None:
            grand_total_with_tax = pr.grand_total_with_tax
        if supplier_name is None:
            supplier_name = pr.supplier_name

    original_barcodes = [line.barcode for line in vision_lines]

    t0 = time.perf_counter()
    match_line_items(vision_lines, reference)
    timing["match_baseline"] = time.perf_counter() - t0

    ref_barcodes = {item.barcode for item in reference if item.barcode}
    enhance_seconds = reporting_seconds = 0.0
    top1_by_index: dict = {}
    barcode_shortcut = [False] * len(vision_lines)

    for idx, (line, orig_bc) in enumerate(zip(vision_lines, original_barcodes)):
        if orig_bc and orig_bc in ref_barcodes:
            barcode_shortcut[idx] = True
            continue  # نفس سلوك الإنتاج - يُترك كما هو، بدون تقييم مرشّحين

        t0 = time.perf_counter()
        top1 = matching_engine.suggest_candidates(line, reference, supplier_name=supplier_name,
                                                     reference_attrs_index=ref_index, top_n=1)
        reporting_seconds += time.perf_counter() - t0
        top1_by_index[idx] = top1[0] if top1 else None

        t0 = time.perf_counter()
        matching_engine.enhance_one(line, reference, supplier_name=supplier_name,
                                     reference_attrs_index=ref_index, allow_semantic=args.with_semantic)
        enhance_seconds += time.perf_counter() - t0

    timing["match_enhance"] = enhance_seconds
    timing["match_total"] = timing["match_baseline"] + enhance_seconds
    timing["reporting_overhead_suggest_candidates"] = reporting_seconds
    timing["pipeline_wall_clock_total"] = timing["pdf_page_load"] + (timing["vision_total"] or 0.0) + timing["match_total"]

    api_cost_this_run = usage_tracker.get_total_spent()

    pairs, missing_truth_idx, extra_extracted_idx = align_lines(vision_lines, original_barcodes, ground_truth)
    records = build_per_line_records(pairs, missing_truth_idx, extra_extracted_idx, vision_lines, ground_truth,
                                      original_barcodes, top1_by_index, barcode_shortcut)
    metrics = compute_metrics(records)

    write_json_report(folder / "benchmark_results.json", args, started, files, timing, api_cost_this_run,
                       gt_cross_check, amnc_survey, records, metrics, tmp_state_dir)
    write_csv_summary(folder / "benchmark_summary.csv", records)
    write_review_package(folder / "benchmark_review_package.json", records, metrics, amnc_survey, gt_cross_check, args, files)
    print_console_summary(metrics, amnc_survey, gt_cross_check, api_cost_this_run, timing)


if __name__ == "__main__":
    main()
